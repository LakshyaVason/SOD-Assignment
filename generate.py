import pandas as pd
from random import shuffle
from collections import defaultdict

# Step 1: Load the data from the Excel file
df = pd.read_excel('Book1.xlsx')

# Step 2: Clean and preprocess the data if necessary
# (Assuming the data is already in the correct format as seen in the image)

# Step 3: Assign shifts randomly while ensuring no repetitions in a two-week period
def assign_shifts(df):
    assignments = []
    employee_count = defaultdict(int)
    dates = df['Date'].tolist()
    employees = df.columns[1:].tolist()
    
    # Prepare availability
    availability = {date: [] for date in dates}
    for index, row in df.iterrows():
        for employee in employees:
            if row[employee] == 'Yes':
                availability[row['Date']].append(employee)
    
    # Assign employees to dates ensuring no repetition within a two-week period
    last_two_weeks = []
    
    for date in dates:
        available_employees = [emp for emp in availability[date] if emp not in last_two_weeks]
        
        if available_employees:  # Check if there are available employees
            shuffle(available_employees)  # Shuffle to ensure randomness
            assigned_employee = available_employees[0]  # Pick the first after shuffling
            assignments.append((date, assigned_employee))
            employee_count[assigned_employee] += 1
            last_two_weeks.append(assigned_employee)
            if len(last_two_weeks) > 10:  # Maintain a rolling list of 10 days (two weeks)
                last_two_weeks.pop(0)
        else:
            assignments.append((date, 'Noone available'))  # Handle no available employee case
    
    return assignments

assignments = assign_shifts(df)

# Step 4: Save the new assignments to an Excel file
output_df = pd.DataFrame(assignments, columns=['Date', 'Assigned Employee'])

# Adjust column width to fit content
output_df.to_excel('shift_assignments.xlsx', index=False)

# Adjust column width to avoid hashtags (done manually or through an external package, not directly supported by pandas)
# You can adjust the column width manually in Excel after generating the file.


import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime

# Load the workbook and select the active worksheet
workbook = openpyxl.load_workbook('shift_assignments.xlsx')
sheet = workbook.active

# Define the fill for past dates
past_date_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Get the current date
current_date = datetime.now()

# Iterate over the rows and apply the fill to past dates
for row in sheet.iter_rows(min_row=2, min_col=1, max_col=1):  # Assuming dates are in column 1 starting from row 2
    for cell in row:
        if cell.value:
            try:
                cell_date = datetime.strptime(str(cell.value), '%Y-%m-%d %H:%M:%S')
                if cell_date < current_date:
                    cell.fill = past_date_fill
            except ValueError:
                # Handle the error if the date format is incorrect
                pass

# Save the workbook
workbook.save('EST_Assignments_colored.xlsx')